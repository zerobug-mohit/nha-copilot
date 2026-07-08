"""Geography / LGD resolution.

Loads lgd_master.xlsx (states, districts, aliases, splits, normalization rules)
into in-memory lookups at startup and resolves free-text place names to LGD
codes. Handles:
  * the 8 normalization rules from the workbook,
  * common aliases / renamings (Orissa -> Odisha, etc.),
  * ambiguous district names shared across states (one clarifying question),
  * post-2011 district splits (surface the pre-split parent),
  * the deliberate data-quality quirk that Maharashtra is spelled MAHARASTRA
    in the TMS/BIS tables, and
  * the brownfield-state asymmetry (TMS has no claims for 7 states; BIS does).
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from app.config import get_settings

logger = logging.getLogger(__name__)

# States whose claims run on their own SHA trust systems and are therefore
# ABSENT from the TMS claims table, but PRESENT in the BIS registry (§4.6).
BROWNFIELD_STATE_NAMES = {
    "rajasthan",
    "maharashtra",
    "karnataka",
    "andhra pradesh",
    "tamil nadu",
    "telangana",
    "west bengal",
}

# Spelling as it actually appears in the TMS/BIS state_name columns (§7.4).
STATE_NAME_IN_DATA = {
    "maharashtra": "MAHARASTRA",  # missing H in the source LGD file
}


@dataclass
class GeoMatch:
    level: str  # "state" | "district"
    lgd_code: int
    name: str  # canonical name
    state_code: int | None = None
    state_name: str | None = None
    is_brownfield: bool = False
    # spelling to use inside a WHERE clause against the data tables
    name_in_data: str | None = None


@dataclass
class GeoResolution:
    """Result of resolving one place token."""

    status: str  # "resolved" | "ambiguous" | "not_found" | "split_parent"
    matches: list[GeoMatch] = field(default_factory=list)
    message: str | None = None  # clarifying question or note

    @property
    def resolved(self) -> GeoMatch | None:
        return self.matches[0] if self.status == "resolved" and self.matches else None


def _norm(text: str) -> str:
    """Apply the workbook's normalization rules (comparison form only)."""
    if text is None:
        return ""
    s = str(text).strip()
    s = re.sub(r"\s+", " ", s)  # collapse internal whitespace
    s = s.lower()
    s = s.replace("&", "and")
    s = re.sub(r"\([^)]*\)", "", s)  # strip parentheticals
    s = s.replace("-", " ")
    s = re.sub(r"\b(sri|shri|saint|st\.?|pt\.?)\b", "", s)  # honorifics
    s = re.sub(r"\b(district|dist\.?|distt\.?)\b", "", s)  # dist. suffix
    s = re.sub(r"[^a-z0-9 ]+", " ", s)  # drop punctuation (?, commas, etc.)
    s = re.sub(r"\s+", " ", s).strip()
    return s


class GeographyResolver:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        # normalized state name/alias -> (code, canonical_name)
        self._state_by_name: dict[str, tuple[int, str]] = {}
        self._state_name_by_code: dict[int, str] = {}
        self._brownfield_codes: set[int] = set()
        # normalized district name -> list of match dicts
        self._district_by_name: dict[str, list[dict]] = {}
        # normalized pre-split parent name -> note
        self._split_parents: dict[str, list[dict]] = {}
        self._loaded = False

    # ---------- loading ----------
    def load(self) -> None:
        if self._loaded:
            return
        if not self.workbook_path.exists():
            raise FileNotFoundError(
                f"LGD master workbook not found at {self.workbook_path}. "
                "Set REFERENCE_DATA_DIR in .env."
            )
        wb = load_workbook(self.workbook_path, data_only=True, read_only=True)
        self._load_states(wb)
        self._load_districts(wb)
        self._load_aliases(wb)
        self._load_splits(wb)
        wb.close()
        self._loaded = True
        logger.info(
            "Geography loaded: %d states, %d district names, %d brownfield",
            len(self._state_name_by_code),
            len(self._district_by_name),
            len(self._brownfield_codes),
        )

    @staticmethod
    def _rows_with_header(ws, header_keys: set[str]):
        """Yield dict rows, auto-detecting the header line (some sheets have a
        title row above the header)."""
        header = None
        for raw in ws.iter_rows(values_only=True):
            cells = [("" if c is None else str(c).strip()) for c in raw]
            if header is None:
                lowered = {c.lower() for c in cells}
                if header_keys & lowered:
                    header = [c.lower() for c in cells]
                continue
            if not any(cells):
                continue
            yield dict(zip(header, cells))

    def _load_states(self, wb) -> None:
        ws = wb["lgd_states"]
        for row in self._rows_with_header(ws, {"lgd_state_code", "lgd_state_name"}):
            code_raw = row.get("lgd_state_code", "")
            name = row.get("lgd_state_name", "")
            if not code_raw or not name:
                continue
            try:
                code = int(str(code_raw).lstrip("0") or "0")
            except ValueError:
                continue
            self._state_name_by_code[code] = name
            self._state_by_name[_norm(name)] = (code, name)
            for alt in str(row.get("common_alternate_names", "")).split(";"):
                alt = alt.strip()
                if alt:
                    self._state_by_name[_norm(alt)] = (code, name)
            if _norm(name) in BROWNFIELD_STATE_NAMES:
                self._brownfield_codes.add(code)

    def _load_districts(self, wb) -> None:
        ws = wb["lgd_districts"]
        for row in self._rows_with_header(
            ws, {"lgd_district_code", "lgd_district_name"}
        ):
            dcode_raw = row.get("lgd_district_code", "")
            dname = row.get("lgd_district_name", "")
            scode_raw = row.get("lgd_state_code", "")
            sname = row.get("lgd_state_name", "")
            if not dcode_raw or not dname or not _norm(dname):
                continue
            try:
                dcode = int(float(dcode_raw))
                scode = int(str(scode_raw).lstrip("0") or "0")
            except ValueError:
                continue
            entry = {
                "district_code": dcode,
                "district_name": dname,
                "state_code": scode,
                "state_name": sname,
            }
            self._district_by_name.setdefault(_norm(dname), []).append(entry)
            # also index the older / census name for traceability
            for alt_key in ("previous_name_2022", "census_2011_district_name"):
                alt = row.get(alt_key, "")
                if alt and _norm(alt) != _norm(dname):
                    self._district_by_name.setdefault(_norm(alt), []).append(entry)

    def _load_aliases(self, wb) -> None:
        if "common_aliases" not in wb.sheetnames:
            return
        ws = wb["common_aliases"]
        for row in self._rows_with_header(ws, {"alias_type", "source_name"}):
            atype = row.get("alias_type", "")
            source = row.get("source_name", "")
            code_raw = row.get("canonical_lgd_code", "")
            canonical = row.get("canonical_name", "")
            if atype == "state" and source and code_raw:
                try:
                    code = int(str(code_raw).lstrip("0") or "0")
                except ValueError:
                    continue
                name = self._state_name_by_code.get(code, canonical)
                self._state_by_name[_norm(source)] = (code, name)

    def _load_splits(self, wb) -> None:
        if "district_splits" not in wb.sheetnames:
            return
        ws = wb["district_splits"]
        for row in self._rows_with_header(ws, {"new_district_name", "parent_district_2011"}):
            parent = row.get("parent_district_2011", "")
            new = row.get("new_district_name", "")
            if not parent:
                continue
            # parent may be "A / B"
            for p in re.split(r"[/,]", parent):
                p = p.strip()
                if p:
                    self._split_parents.setdefault(_norm(p), []).append(
                        {"new_district": new, "state": row.get("state_name", "")}
                    )

    # ---------- resolution ----------
    def resolve(self, place: str) -> GeoResolution:
        self.load()
        key = _norm(place)
        if not key:
            return GeoResolution(status="not_found")

        # 1. state?
        if key in self._state_by_name:
            code, name = self._state_by_name[key]
            return GeoResolution(
                status="resolved",
                matches=[self._state_match(code, name)],
            )

        # 2. district?
        if key in self._district_by_name:
            entries = self._district_by_name[key]
            unique_states = {e["state_code"] for e in entries}
            if len(unique_states) > 1:
                opts = ", ".join(
                    sorted({f"{e['district_name']} in {e['state_name']}" for e in entries})
                )
                return GeoResolution(
                    status="ambiguous",
                    matches=[self._district_match(e) for e in entries],
                    message=f"There are multiple districts named "
                    f"'{place.strip()}' — {opts}. Which did you mean?",
                )
            return GeoResolution(
                status="resolved", matches=[self._district_match(entries[0])]
            )

        # 3. pre-split parent name?
        if key in self._split_parents:
            children = self._split_parents[key]
            names = ", ".join(sorted({c["new_district"] for c in children if c["new_district"]}))
            return GeoResolution(
                status="split_parent",
                message=(
                    f"'{place.strip()}' appears to be a pre-2011 parent district that "
                    f"has since been split into: {names}. Please query one of these, "
                    "or ask for the parent region explicitly."
                ),
            )

        return GeoResolution(status="not_found")

    def _state_match(self, code: int, name: str) -> GeoMatch:
        norm = _norm(name)
        return GeoMatch(
            level="state",
            lgd_code=code,
            name=name,
            state_code=code,
            state_name=name,
            is_brownfield=code in self._brownfield_codes or norm in BROWNFIELD_STATE_NAMES,
            name_in_data=STATE_NAME_IN_DATA.get(norm, name.upper()),
        )

    def _district_match(self, entry: dict) -> GeoMatch:
        snorm = _norm(entry["state_name"])
        return GeoMatch(
            level="district",
            lgd_code=entry["district_code"],
            name=entry["district_name"],
            state_code=entry["state_code"],
            state_name=entry["state_name"],
            is_brownfield=entry["state_code"] in self._brownfield_codes
            or snorm in BROWNFIELD_STATE_NAMES,
            name_in_data=entry["district_name"].upper(),
        )

    def detect(self, text: str) -> list[GeoResolution]:
        """Scan free text for known state/district names and resolve them.

        States take precedence over districts on an exact phrase hit. Returns one
        GeoResolution per distinct place found (including any ambiguous ones).
        """
        self.load()
        norm_text = f" {_norm(text)} "
        results: list[GeoResolution] = []
        seen: set[str] = set()

        # States first (fewer, higher confidence).
        for key, (code, name) in self._state_by_name.items():
            if key and f" {key} " in norm_text and name not in seen:
                seen.add(name)
                results.append(
                    GeoResolution(status="resolved", matches=[self._state_match(code, name)])
                )

        # Districts (skip if same token already matched a state).
        for key, entries in self._district_by_name.items():
            if not key or f" {key} " not in norm_text:
                continue
            label = entries[0]["district_name"]
            if label in seen:
                continue
            seen.add(label)
            res = self.resolve(label)
            if res.status in ("resolved", "ambiguous"):
                results.append(res)
        return results

    def is_brownfield_state_code(self, code: int) -> bool:
        self.load()
        return code in self._brownfield_codes


_resolver: GeographyResolver | None = None


def get_geography() -> GeographyResolver:
    global _resolver
    if _resolver is None:
        settings = get_settings()
        _resolver = GeographyResolver(settings.reference_dir / "lgd_master.xlsx")
    return _resolver
